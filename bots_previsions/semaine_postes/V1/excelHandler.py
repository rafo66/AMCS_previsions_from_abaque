import datetime
import os, os.path
import pandas as pd
import time

'Excel formatter dependecy'
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import sys

        
#WOIPPY data : 
oldLines =     ["D10", "D10R11", "D14R11", "D20", "D7R10", "FIMI", "FIMIR3B", "FIN","IOWA", "L1", "LAS1.", "P3", "R1", "R10", "R11", "R2", "R3B", "R2B", "R6", "R7", "P3R1", "P3R6"]
newLines =     ["L1", "L1",      "L1",      "L1", "L1",     "L1",  "L1",      "L1","L1", "L1", "LASS1,", "P3", "R1", "R1", "R6", "R6", "R6", "R6","R6", "R1", "P3", "P3"]
avgLineProto = [4.15, 4.15,      4.15,      4.15, 4.15,     4.15, 4.15,       4.15,4.15, 4.15, 0.39,     2.25, 14,   14,   15.8, 15.8, 15.8, 15.8, 15.8, 14, 2.25, 2.25]
avgLineSerie = [6.96, 6.69,      6.69,      6.96, 6.96,     6.96, 6.96,       6.96,6.96,  6.96,  0.7,      2.5, 23.2, 23.2, 39,   39,   39,   39,39,   23.2,  2.5,  2.5]




VERSION = 0.1
CURENT_TIME_ZONE = "Europe/Paris"


class MatchingProductivities:
    '''
        Takes 2 dataframes as input :
        - exceptionReportDf : dataframe of the exception report, with only Woippy plant data
        - abaqueDf : dataframe of the abaque, with all the data

        Returns the exceptionReportDf with a new column "Productivity" filled with the matched productivity for each row, and a new column "Match Level" filled with the level of the match (0 to 4, or -1 if no match)

        Exemple : 
        
        mP = MatchingProductivities(self.getExceptionReportDf(), self.getAbaqueDf(), self.cacheFile)
        self.newExceptionReportDf = mP.exceptionReportDf
        
        if cacheFile is empty or None, no caching will be done. Else articleCachedProductivities.txt is used

        Dependencies : 
            - OS access to read and write cache file
            - pandas for dataframe manipulation


        

        ' LEVEL 0: Exact Match Article            
        ' LEVEL 1: Exact Match (Name + Proto + 3 Dims)
        ' LEVEL 2: Partial Match (Name + Proto + Dim1)
        ' LEVEL 3: Partial Match (Name + Proto)
        ' LEVEL 4: Moyenne ligne

    '''
            
    def __init__(self, exceptionReportDf, abaqueDf, cacheFile="articleCachedProductivities.txt"):
        self.exceptionReportDf = exceptionReportDf
        self.abaqueDf = abaqueDf
        self.cacheFile = cacheFile

        try:
            self.articleCachedProductivities = self.loadArticleCachedProductivities()
        except Exception as e:
            printerUtil("Error while loading cached productivities: ", e)
            self.articleCachedProductivities = {}

        self.matchProductivities()
 
    def loadArticleCachedProductivities(self):
        if self.cacheFile == None or self.cacheFile == "":
            return {}
        
        # Load cached productivities from a file if it exists, otherwise return an empty dictionary
        if os.path.exists(self.cacheFile):
            with open(self.cacheFile, "r") as f:
                lines = f.readlines()
                articleCachedProductivities = {}
                for line in lines:
                    article, prod, details, level = line.strip().split(":")
                    articleCachedProductivities[int(article)] = [float(prod), details, int(level)]
                printerUtil("Loaded cached productivities from file ")
                return articleCachedProductivities
        else:
            return {}
        
    def saveCachedProductivities(self):
        if self.cacheFile == None or self.cacheFile == "":
            return
        
        
        # Save cached productivities to a file
        with open(self.cacheFile, "w") as f:
            for article, prod in self.articleCachedProductivities.items():
                f.write(f"{int(article)}:{round(float(prod[0]), 2)}:{prod[1]}:{prod[2]}\n")

    def matchProductivities(self):
        '''
        Exception Report Columns: Sales Document | Material | Plant | Last Updated 
        | Sold-to party | Name of sold-to party | Price Partner | Name of Price partner 
        | Ship-to Party | Name of ship-to party | Ship-To Location | External office code 
        | Name of External Office | Back Office code | Name of Back Office | Sales Type 
        | Sales unit | Numerator | Denominator | Number of Pieces | Call-Offs | Sold to Article Number 
        | Aramis Grade | Thickness | Width | Length | Base Unit of Measure | Customer Order Reference 
        | Internal Order Number | Material for BOM | Critical Part | Mill | Leverage contract list 
        | Leverage Article list | BOM Alert | Routing | Backlog | Forecast W | Forecast W1 | Forecast W2 
        | Forecast W3 | Forecast W4 | Forecast W5 | Forecast W6 | Forecast W7 | Forecast W8 | Forecast W-1 
        | Forecast W-2 | Aramis Open Quantity_W | Aramis Open Quantity_W1 | Aramis Open Quantity_W2 
        | Aramis Open Quantity_W3 | Aramis Open Quantity_W4 | Aramis Open Quantity_W5 | Aramis Open Quantity_W6 
        | Aramis Open Quantity_W7 | Aramis Open Quantity_W8 | Aramis Open Quantity_D0 | Aramis Open Quantity_D1 
        | Aramis Open Quantity_D2 | Aramis Open Quantity_D3 | Aramis Open Quantity_D4 | Aramis Open Quantity_D5 
        | Aramis Open Quantity_D6 | Aramis Open Quantity_D7 | Aramis Open Quantity_D8 | Aramis Open Quantity_D9 
        | Aramis Open Quantity_D10 | Aramis Open Quantity_D11 | Aramis Open Quantity_D12 | Aramis Open Quantity_D13 
        | Aramis Open Quantity_D14 | Aramis Open Quantity_D15 | Call-Offs_W | Call-Offs_W1 | Call-Offs_W2 | Call-Offs_W3 
        | Call-Offs_W4 | STOCK_FG_FREE | Blocked Quality | Qual Inspection | Consignment Stock | On a transport request 
        | Unrestricted RM | Stock in tfr RM | Quality Inspection RM | Restricted-Use Stock RM | Blocked RM 
        | Unrestr. Consignment RM | Cnsgt in Inspection RM | Restr. Consignment RM | Blocked Consignment RM 
        | RTS at the Mill | RTS at the Mill UOM | In Transit from the Mill | In Transit from the Mill UOM 
        | Backlog Detail W-0 | Backlog Detail W-1 | Backlog Detail W-2 | Backlog Detail W-3 | Backlog Detail W-4 
        | Backlog Detail W-5 | Backlog Detail W-6 | Backlog Detail W-7 | Backlog Detail W-8 | Backlog Details W-9 and more 
        | Backlog Detail D-1 | Backlog Detail D-2 | Backlog Detail D-3 | Backlog Detail D-4 | Backlog Detail D-5 
        | Backlog Detail D-6 | Backlog Detail D-7 | Production : Unprocessed quantity past | Production : Unprocessed quantity W0 
        | Production : Unprocessed quantity W1 | Production : Unprocessed quantity W2 | Production : Unprocessed quantity W3 
        | Production : unplanned quantity | Production : Finish planned in the Past | Finish Planned W-0 | Finish Planned W-1 
        | Finish Planned W-2 | Finish Planned W-3 | Finish Planned W-4 | Unit | Date of the last Call-off Cust Version 
        | Date of the last forecast Cust version | Last Updated week | To be delivered W-0 | To be delivered W-1 
        | To be delivered W-2 | To be delivered W-3 | To be delivered W-4 | To be delivered W-5 | To be delivered W-6 
        | To be delivered W-7 | To be delivered W-8 | Quantity in transit to customer | Total Stock | To be delivered D-0 
        | To be delievred D-1 | To be delievred D-2 | To be delievred D-3 | To be delievred D-4 | To be delievred D-5 
        | To be delievred D-6 | To be delievred D-7 | To be delievred D-8 | RM | Plant_RM | Nb Sales / RM | FG (Total) 
        | WO Plan | WO Non Plan (SO) | WO Non Plan (RM) | RM (Plant + Consi) | RM Available | RM W | RM W+1 | RM W+2 
        | RTB Ship Backlog | RTB Ship W | RTB Ship W+1 | RTB Ship W+2 | RTB Ship W+3 | RTB Ship W+4 | RTB Ship W+5 
        | RTB Ship W+6 | RTB Ship W+7 | RTB Ship W+8 | Lack FG Backlog (SO) | Lack FG W (SO) | Lack FG W+1 (SO) 
        | Lack FG W+2 (SO) | Lack FG+Plan Backlog (SO) | Lack FG+Plan W (SO) | Lack FG+Plan W+1 (SO) | Lack FG+Plan W+2 (SO) 
        | Lack FG+Plan Backlog (RM) | Lack FG+Plan W (RM) | Lack FG+Plan W+1 (RM) | Lack FG+Plan W+2 (RM) 
        | Plan Order Backlog (SO) | Plan Order W (SO) | Plan Order W+1 (SO) | Plan Order W+2 (SO) | Nb Plan Order Backlog (SO) 
        | Nb Plan Order W (SO) | Nb Plan Order W+1 (SO) | Nb Plan Order W+2 (SO) | Plan Order Backlog (RM) | Plan Order W (RM) 
        | Plan Order W+1 (RM) | Plan Order W+2 (RM) | Lack RM Plant Backlog (SO) | Lack RM Plant W (SO) | Lack RM Plant W+1 (SO) 
        | Lack RM Plant W+2 (SO) | Lack RM Plant Backlog (SO).1 | Lack RM Plant W Cumul (SO) | Lack RM Plant W+1 Cumul (SO) 
        | Lack RM Plant W+2 Cumul (SO) | Lack RM Plant Backlog (RM) | Lack RM Plant W (RM) | Lack RM Plant W+1 (RM) 
        | Lack RM Plant W+2 (RM) | Lack RM Plant Backlog (RM).1 | Lack RM Plant W Cumul (RM) | Lack RM Plant W+1 Cumul (RM) 
        | Lack RM Plant W+2 Cumul (RM) | Lack RM AM Backlog | Lack RM AM W | Lack RM AM W+1 | Lack RM AM W+2 | FG Backlog 
        | FG W | FG W+1 | FG W+2 | RM Plant Backlog | RM Plant W | RM Plant W+1 | RM Plant W+2 | RM AM Backlog | RM AM W 
        | RM AM W+1 | RM AM W+2 | Alerte Backlog | Alerte W | Alerte W+1 | Alerte W+2 | GOR Backlog | GOR W | GOR W+1 | GOR W+2 
        | Lack RM Type | Lack RM Cover | GOR TYPE Backlog | GOR TYPE W | GOR TYPE W+1 | GOR TYPE W+2 | Horizon Backlog 
        | Horizon W | Horizon W+1 | Horizon W+2 | No Needs Horizon W+2 | Action to do | Top 10 | Stay to deliver following 2 weeks 
        | total ship | total cut back & week | total call Back & week | FG Available for week ongoing | campagne optimizer  
        | Weekly Coverage FG | Weekly Coverage RM | RM.1 | Weekly Consumption | RTB Ship Cumul Backlog | RTB Ship Cumul W 
        | RTB Ship Cumul W+1 | RTB Ship Cumul W+2 | RTB Ship Cumul W+3 | RTB Ship Cumul W+4 | RTB Ship Cumul W+5 
        | RTB Ship Cumul W+6 | RTB Ship Cumul W+7 | RTB Ship Cumul W+8 | FG | FG Backlog.1 | FG W.1 | FG W+1.1 | FG W+2.1 
        | FG W+3 | FG W+4 | FG W+5 | FG W+6 | FG W+7 | FG W+8 | Plan | Plan Backlog | Plan W | Plan W+1 | Plan W+2 | Plan W+3 
        | Plan W+4 | Plan W+5 | Plan W+6 | Plan W+7 | Plan W+8 | WO | WO Backlog | WO W | WO W+1 | WO W+2 | WO W+3 | WO W+4 
        | WO W+5 | WO W+6 | WO W+7 | WO W+8 | Plan Order Backlog | Plan Order W | Plan Order W+1 | Plan Order W+2 | Plan Order W+3 
        | Plan Order W+4 | Plan Order W+5 | Plan Order W+6 | Plan Order W+7 | Plan Order W+8 | NB PlOr Backlog | NB PlOr W 
        | NB PlOr W+1 | NB PlOr W+2 | NB PlOr W+3 | NB PlOr W+4 | NB PlOr W+5 | NB PlOr W+6 | NB PlOr W+7 | NB PlOr W+8 | RM.2 
        | RM Backlog | RM W.1 | RM W+1.1 | RM W+2.1 | RM W+3 | RM W+4 | RM W+5 | RM W+6 | RM W+7 | RM W+8 | Bes Transit Backlog 
        | Bes Transit W | Bes Transit W+1 | Bes Transit W+2 | Bes Transit W+3 | Bes Transit W+4 | Bes Transit W+5 | Bes Transit W+6 
        | Bes Transit W+7 | Bes Transit W+8 | NB Transit Backlog | NB Transit W | NB Transit W+1 | NB Transit W+2 | NB Transit W+3 
        | NB Transit W+4 | NB Transit W+5 | NB Transit W+6 | NB Transit W+7 | NB Transit W+8 | Transit | Transit Backlog | Transit W 
        | Transit W+1 | Transit W+2 | Transit W+3 | Transit W+4 | Transit W+5 | Transit W+6 | Transit W+7 | Transit W+8 | Bes RTS Backlog 
        | Bes RTS W | Bes RTS W+1 | Bes RTS W+2 | Bes RTS W+3 | Bes RTS W+4 | Bes RTS W+5 | Bes RTS W+6 | Bes RTS W+7 | Bes RTS W+8 
        | NB RTS Backlog | NB RTS W | NB RTS W+1 | NB RTS W+2 | NB RTS W+3 | NB RTS W+4 | NB RTS W+5 | NB RTS W+6 | NB RTS W+7 | NB RTS W+8 
        | RTS | RTS Backlog | RTS W | RTS W+1 | RTS W+2 | RTS W+3 | RTS W+4 | RTS W+5 | RTS W+6 | RTS W+7 | RTS W+8 | NO RM Backlog | NO RM W 
        | NO RM W+1 | NO RM W+2 | NO RM W+3 | NO RM W+4 | NO RM W+5 | NO RM W+6 | NO RM W+7 | NO RM W+8 | Coverage Backlog | Coverage W 
        | Coverage W+1 | Coverage W+2 | Coverage W+3 | Coverage W+4 | Coverage W+5 | Coverage W+6 | Coverage W+7 | Coverage W+8 
        | Perf délais Cumul forecast W-2 + W-1 | Perf délais Cumul forecast W-2 to W+2 | Perf délais  Cumul forecast W-2+ W+8 
        | Perf délais Backlog sup. 0,5 T | Analyse Backlog W-1 | Analyse Forecast w-1 | Transit on FG | Week No Covered by FG 
        | Flag Coverage FG | Week No Covered by Stock SSC | Flag Coverage Stock SSC | RTS Horizon W+8 | Concat PLANT RM | yy | 

        '''
        '''
        Abaque Columns: OF | Prod T/h/OF | Bandes (/OF) | Scindage (/OF) | Date 1 
        | Temps 1 (h) | Année 1 | Tonnes | heures | Épaisseur Nominal | Largeur 
        | Longueur | Proto | Poste | Nb pieces | Date 2 | Temps 2 (h) | Année 2 
        | Clients | Articles |
        '''

        
        # add a column to exceptionReportDf called "Productivity" and fill it with -1
        self.exceptionReportDf["Productivity"] = float(-1)
        
        # add a column named Abaque Indexes that stores abaques Indexes
        self.exceptionReportDf["Abaque Indexes"] = ""


        subIndex=0
        for index, row in self.exceptionReportDf.iterrows():
            prod, details, level = self.getProductivityForRow(index, row)
            self.exceptionReportDf.at[index, "Productivity"] = round(float(prod), 2)
            self.exceptionReportDf.at[index, "Abaque Indexes"] = str(details) + "#" + str(level)

            if subIndex % 50 == 0:
                printerUtil(f"Row {subIndex} / {index} / {len(self.exceptionReportDf)} level: {level}, Prod: {round(float(prod), 2)}")
                try:
                    self.saveCachedProductivities()
                except Exception as e:
                    printerUtil("Error while saving cached productivities: ", e)

            subIndex+=1

        return self.exceptionReportDf

    def getProductivityForRow(self, index, row):
        articleReport = int(row["Material"])
        clientReport = row["Name of sold-to party"]
        salesTypeReport = row["Sales Type"]
        
        dim1Report = row["Length"]
        dim2Report = row["Width"]
        dim3Report = row["Thickness"]
        dim1Report = round(float(dim1Report), 2)
        dim2Report = round(float(dim2Report), 2)
        dim3Report = round(float(dim3Report), 2)

        self.curentDetails = []
        self.curentProductivity = -1

        if articleReport in self.articleCachedProductivities:
            return self.articleCachedProductivities[articleReport][0], self.articleCachedProductivities[articleReport][1], self.articleCachedProductivities[articleReport][2]



        self.filterLevel0(articleReport)
        if self.curentProductivity != -1:
            self.articleCachedProductivities[articleReport] = [round(float(self.curentProductivity), 2), self.curentDetails, 0]
            return round(float(self.curentProductivity), 2), self.curentDetails, 0
        

        
        self.filterLevel1(clientReport, salesTypeReport, dim1Report, dim2Report, dim3Report)
        if self.curentProductivity != -1:
            self.articleCachedProductivities[articleReport] = [round(float(self.curentProductivity), 2), self.curentDetails, 1]
            return round(float(self.curentProductivity), 2), self.curentDetails, 1
        
        
        self.filterLevel2(clientReport, salesTypeReport, dim3Report)
        if self.curentProductivity != -1:
            self.articleCachedProductivities[articleReport] = [round(float(self.curentProductivity), 2), self.curentDetails, 2]
            return round(float(self.curentProductivity), 2), self.curentDetails, 2
        

        self.filterLevel3(clientReport, salesTypeReport)
        if self.curentProductivity != -1:
            self.articleCachedProductivities[articleReport] = [round(float(self.curentProductivity), 2), self.curentDetails, 3]
            return round(float(self.curentProductivity), 2), self.curentDetails, 3
        
        
        oldLineReport = row["Routing"]
        self.filterLevel4(oldLineReport, salesTypeReport)
        if self.curentProductivity != -1:
            self.articleCachedProductivities[articleReport] = [round(float(self.curentProductivity), 2), self.curentDetails, 4]
            return round(float(self.curentProductivity), 2), self.curentDetails, 4
        


        printerUtil("Absolutly no match for row ", index, articleReport, clientReport, salesTypeReport, dim1Report, dim2Report, dim3Report, oldLineReport)
        self.articleCachedProductivities[articleReport] = [-1, self.curentDetails, -1]
        return -1, self.curentDetails, -1
        
    def isProtoName(self, protoReport):
        if protoReport == "AM Prototype Order" or protoReport == "AM Free Prototype":
            return 'VRAI'
        return 'FAUX'

    def filterLevel0(self, article):
        potentialProd = []
        self.curentDetails = []

        for index, row in self.abaqueDf.iterrows():
            try:
                if str(int(article)) in str(row["Articles"]):
                    potentialProd.append(row["Prod T/h/OF"])
                    self.curentDetails.append(index)
            except:
                pass

        if len(potentialProd) > 0:
            self.curentProductivity = sum(potentialProd)/len(potentialProd)
        else:
            self.curentProductivity = - 1
    
    def filterLevel1(self, name, proto, dim1, dim2, dim3):
        potentialProd = []
        self.curentDetails = []

        for index, row in self.abaqueDf.iterrows():
            try:
                b1 = str(name) in str(row["Clients"])
                
                if self.isProtoName(proto) == 'VRAI':
                    b2 = int(row["Proto"]) == 1
                else:
                    b2 = int(row["Proto"]) == 0

                
                b3 = dim1 == round(float(row["Épaisseur Nominal"]), 2)
                b4 = dim2 == round(float(row["Largeur"]), 2)
                b5 = dim3 == round(float(row["Longueur"]), 2)

                if b1 and b2 and b3 and b4 and b5:
                    potentialProd.append(row["Prod T/h/OF"])
                    self.curentDetails.append(index)
            except:
                pass

        if len(potentialProd) > 0:
            self.curentProductivity = sum(potentialProd)/len(potentialProd)
        else:
            self.curentProductivity = -1
           
    def filterLevel2(self, name, proto, dim1):
        potentialProd = []
        self.curentDetails = []

        for index, row in self.abaqueDf.iterrows():
            try:
                b1 = str(name) in str(row["Clients"])
                
                if self.isProtoName(proto) == 'VRAI':
                    b2 = int(row["Proto"]) == 1
                else:
                    b2 = int(row["Proto"]) == 0

                
                b3 = dim1 == round(float(row["Épaisseur Nominal"]), 2)


                if b1 and b2 and b3:
                    potentialProd.append(row["Prod T/h/OF"])
                    self.curentDetails.append(index)
            except:
                pass

        if len(potentialProd) > 0:
            self.curentProductivity = sum(potentialProd)/len(potentialProd)
        else:
            self.curentProductivity = -1
    
    def filterLevel3(self, name, proto):
        potentialProd = []
        self.curentDetails = []

        for index, row in self.abaqueDf.iterrows():
            try:
                b1 = str(name) in str(row["Clients"])
                
                if self.isProtoName(proto) == 'VRAI':
                    b2 = int(row["Proto"]) == 1
                else:
                    b2 = int(row["Proto"]) == 0

                

                if b1 and b2:
                    potentialProd.append(row["Prod T/h/OF"])
                    self.curentDetails.append(index)
            except:
                pass

        if len(potentialProd) > 0:
            self.curentProductivity = sum(potentialProd)/len(potentialProd)
        else:
            self.curentProductivity = -1
        
    def filterLevel4(self, oldLine, proto):
        self.curentDetails = []

        if self.isProtoName(proto) == 'VRAI':
            if oldLine in oldLines:
                self.curentDetails = "Average of line " + str(oldLine) + " for proto articles in abaque"
                self.curentProductivity = avgLineProto[oldLines.index(oldLine)]
                return
        else:
            if oldLine in oldLines:
                self.curentDetails = "Average of line " + str(oldLine) + " for serie articles in abaque"
                self.curentProductivity = avgLineSerie[oldLines.index(oldLine)]
                return
            
        self.curentDetails = "No average line data for " + str(oldLine) + " proto" + str(proto) 
        self.curentProductivity = -1
          
class excelHandler:
    '''
    Opens and get dataframe from Exception_report.xlsb
        => Filters the dataframe to only get Woippy plant data
        => Export dataframe 
    Opens and get dataframe from Abaque.xlsm
        => Export dataframe

    Run MatchingProductivities function to get a productivity for each client in report



    
    '''
    def __init__(self, exceptionReportPath, abaquePath, plantName, cacheFile="articleCachedProductivities.txt", bypassCalculs=False, DF_cacheFile="Processed_Exception_report.xlsx"):
        self.exceptionReportPath = exceptionReportPath
        self.abaquePath = abaquePath
        self.plantName = plantName
        self.cacheFile = cacheFile
        self.DF_cacheFile = DF_cacheFile
        self.infoText = ""

        self.bypassCalculs = bypassCalculs
        self.abaqueDF = self.getAbaqueDf()
        try:
            self.filterAbaque()
        except Exception as e:
            printerUtil("Error filtering abaque: ", e)
            exit(1)


        if bypassCalculs:
            self.newDf = pd.read_excel(self.DF_cacheFile)
            self.exceptionReportLastModified = os.path.getmtime(self.exceptionReportPath)    

            #self.build_details()
        else:
            self.exceptionReportDF = self.getExceptionReportDf()

            try:
                self.filterFalseBacklog()
            except Exception as e:
                printerUtil("Error filtering false backlog: ", e)
                exit(1)

            mP = MatchingProductivities(self.exceptionReportDF, self.abaqueDF, self.cacheFile)
            self.newExceptionReportDf = mP.exceptionReportDf
    
            self.processExceptionReport()


    def getExceptionReportDf(self):
        #save lastModified date of exceptionReportPath in self.exceptionReportLastModified
        self.exceptionReportLastModified = os.path.getmtime(self.exceptionReportPath)
        try:
            sheetName = "Sheet1"
            df = pd.read_excel(self.exceptionReportPath, sheet_name=sheetName)
            newDf = df[df["Plant"].str.contains(self.plantName, case=False, na=False)]
            return newDf
        except Exception as e:
            printerUtil("Error while reading or filtering exception report : ", e)
            exit(1)

    def filterFalseBacklog(self):
        # Exclude some backlog lines:
        # If Sales Type contains "Hire Work" AND Backlog > 100 AND sum(Forecast W..W8) == 0 => set Backlog to 0
        forecast_cols = [f"Forecast W{i}" if i else "Forecast W" for i in range(0, 9)]

        # Ensure numeric (NaN -> 0) for the computation
        for c in forecast_cols + ["Backlog"]:
            if c in self.exceptionReportDF.columns:
                self.exceptionReportDF[c] = pd.to_numeric(self.exceptionReportDF[c], errors="coerce").fillna(0)
            else:
                self.infoText += f"Column '{c}' not found in exception report. \n"
                self.exceptionReportDF[c] = 0  # Add the column with default 0 if it's missing
                printerUtil(f"Column '{c}' not found in exception report. Added with default 0.")

        sales_type = self.exceptionReportDF["Sales Type"].astype(str)
        forecast_sum = self.exceptionReportDF[forecast_cols].sum(axis=1)

        mask = (
            sales_type.str.contains("Hire Work", case=False, na=False)
            & (self.exceptionReportDF["Backlog"] > 20)
            & (forecast_sum == 0)
        )

        excluded = int(mask.sum())
        if excluded:
            printerUtil(f"Excluding backlog for {excluded} 'Hire Work' lines (Backlog>20 and Forecast sum=0).")
            self.exceptionReportDF.loc[mask, "Backlog"] = 0
            printerUtil("Number of lines with Backlog > 0 after cleanup: ", (self.exceptionReportDF["Backlog"] > 0).sum())
            # if LAS1 in Routing, replace by LAS1.
            self.exceptionReportDF["Routing"] = self.exceptionReportDF.apply(lambda row: "LAS1." if "LAS1" in str(row["Routing"]) else row["Routing"], axis=1)



    def getAbaqueDf(self):
        try:
            t1 = time.time()
            df = pd.read_excel(self.abaquePath)        
            return df
        except Exception as e:
            printerUtil("Error while reading abaque : ", e)
            exit(1)

    def filterAbaque(self):
        # if poste = LAS1 keep only the lines where 2026 is in Année 1 
        self.abaqueDF = self.abaqueDF[~((self.abaqueDF["Poste"] == "LAS1") & (self.abaqueDF["Année 1"] != 2026))]

    def processExceptionReport(self):
        '''
            Create Postes columns, and format routing and proto
        '''

        intressingFields = ["Name of sold-to party", "Abaque Indexes", "Routing", "Material", "Productivity", "Forecast W", "Forecast W1", "Forecast W2", "Forecast W3", "Forecast W4", "Forecast W5", "Forecast W6", "Forecast W7", "Forecast W8", "Backlog", "Thickness", "Width", "Length", "STOCK_FG_FREE"]
    
        calculatedFields = ["Forecast W", "Forecast W1", "Forecast W2", "Forecast W3", "Forecast W4", "Forecast W5", "Forecast W6", "Forecast W7", "Forecast W8", "Backlog"]
        for field in calculatedFields:
            newFieldName = field + " (Postes)"
            self.newExceptionReportDf[newFieldName] = self.newExceptionReportDf.apply(lambda row: round(float(row[field]) / float(row["Productivity"]) / 7.5, 2) if row["Productivity"] != -1 and float(row["Productivity"]) != 0 else -1, axis=1)

        # apply newLine to routing column and save it in a new column called "Routing (Postes)"
        self.newExceptionReportDf["New Routing"] = self.newExceptionReportDf.apply(lambda row: newLines[oldLines.index(row["Routing"])] if row["Routing"] in oldLines else row["Routing"], axis=1)
        # apply isProtoName to proto column and save it in a new column called "Is Proto"
        self.newExceptionReportDf["Is Proto"] = self.newExceptionReportDf.apply(lambda row : self.isProtoName(row["Sales Type"]), axis=1)

                                                                                
        newDf = self.newExceptionReportDf[intressingFields + [field + " (Postes)" for field in calculatedFields] + ["New Routing", "Is Proto"]]
        

        printerUtil("Number of lines before cleanup: ", len(newDf))
        newDf = newDf[newDf.apply(lambda row: all(self.isValid(row[field + " (Postes)"]) for field in calculatedFields), axis=1)]
        newDf = newDf[newDf.apply(lambda row: sum(float(row[field]) for field in calculatedFields) != 0, axis=1)]
        printerUtil("Number of lines after cleanup: ", len(newDf))




        # Create a table with each routing, and subdividing it in proto or not, and 1 column Sum of forecast W
        
        newDf.to_excel(self.DF_cacheFile, index=False)
        self.newDf = newDf

    def get_newDf(self):
        return self.newDf
    
    def isProtoName(self, protoReport):
        if protoReport == "AM Prototype Order" or protoReport == "AM Free Prototype":
            return 'VRAI'
        return 'FAUX'

    def isValid(self, value):
        if value == -1 or value == None or value == "" or str(value).lower() == "nan":
            return False
        return True

class outputFormatter:
    def __init__(self, df, abaqueDF, tp=r"OutputTemplate.xlsm", exceptionReportLastModified=None):
        self.df = df
        self.abaqueDF = abaqueDF
        self.exceptionReportLastModified = exceptionReportLastModified
        
        self.outputExcel(template_path=tp)

    def outputExcel(self, template_path):
        
        try:
            wb = load_workbook(template_path, keep_vba=True)
            ws = wb["Details"]
            ws.delete_rows(1, ws.max_row)
            ws = wb["Resultats"]
        except Exception as e:
            printerUtil("Error while loading Excel template: ", e)
            exit(1)


        start_row = 5  # L1 total starts here
        current_row = start_row






        columnsToSum = ["Backlog (Postes)", "Backlog", "Forecast W (Postes)", "Forecast W", "Forecast W1 (Postes)", "Forecast W1", "Forecast W2 (Postes)", "Forecast W2", "Forecast W3 (Postes)", "Forecast W3", "Forecast W4 (Postes)", "Forecast W4", "Forecast W5 (Postes)", "Forecast W5", "Forecast W6 (Postes)", "Forecast W6", "Forecast W7 (Postes)", "Forecast W7", "Forecast W8 (Postes)", "Forecast W8"]
        summary = self.df.groupby(["New Routing", "Is Proto"]).agg({col: "sum" for col in columnsToSum}).reset_index()

        self.detailIndex = 0
        self.curentDetailRow = 2

        self.needsNewDesc = True
        self.lastDetailRow = -1




        # Write summary data to Excel
        for index, row in summary.iterrows():
            for i, col in enumerate(columnsToSum):

                
                
                cell = ws.cell(row=current_row, column=2 + i)

                
                try:
                    value = round(float(row[col]), 2)
                    cell.value = value
                    cell.font = Font(color="000000")
                except Exception as e:
                    printerUtil("Error setting value for cell:", e, " value: ", value, " routing: ", row["New Routing"], " proto: ", row["Is Proto"], " col: ", col)
                    
                    continue



                # Coefficient
                isPostes = True if "(Postes)" in col else False
                if isPostes:
                    cell.value = "=" + str(value) + "/D49"
               


                tonnesColName = col.replace(" (Postes)", "") if isPostes else col 
                postesColName = col if isPostes else col + " (Postes)"
                if row[tonnesColName] == 0 and row[postesColName] == 0:
                    cell.hyperlink = None
                    continue


                self.needsNewDesc = "Postes" in col 

                # Hyperlien interne vers une cellule précise
                if self.needsNewDesc:
                    cell.hyperlink = f"#Details!C{self.curentDetailRow}"
                else:
                    cell.hyperlink = f"#Details!C{self.lastDetailRow}"
                self.lastDetailRow = self.curentDetailRow
                
                

                rootingIndex = newLines.index(row["New Routing"]) if row["New Routing"] in newLines else -1
                protoIndex = 1 if row["Is Proto"] == 'VRAI' else 0
                if self.needsNewDesc:
                    self.createDetail(rootingIndex, protoIndex, isPostes, col, wb)



                # go back to Resultats sheet
                ws = wb["Resultats"]

            
            if current_row %3 == 0:
                current_row=current_row+1
            current_row += 1


        
        current_row = 19

        # write total in column B for backlog postes, and in column C for backlog, and so on for each forecast W
        for i, col in enumerate(columnsToSum):
            isPostes = True if "(Postes)" in col else False
            if isPostes:
                ws.cell(row=current_row, column=2 + i).value = "=" + str(round(float(summary[col].sum()), 2)) + "/D49"
            else:
                ws.cell(row=current_row, column=2 + i).value = round(float(summary[col].sum()), 2)

        # write subTotal in line 4, 7, 10, 13, 16, 19, 22, 25, 28, 31 in column B for backlog postes, and in column C for backlog, and so on for each forecast W
        for i, col in enumerate(columnsToSum):
            isPostes = True if "(Postes)" in col else False

            rows_and_routings = [
            (4, "L1"),
            (7, "LASS1,"),
            (10, "P3"),
            (13, "R1"),
            (16, "R6"),
            ]

            for row, routing in rows_and_routings:
                try:
                    value = round(float(summary[summary["New Routing"] == routing][col].sum()), 2)
                except KeyError:
                    value = 0
                
                
                cell = ws.cell(row=row, column=2 + i)
                
                if isPostes:
                    cell.value = "=" + str(value) + "/D49"
                else:
                    cell.value = value

        # STOCK_FG_FREE
        for forecast_i in range(9):
            forecastText = "Forecast W" if forecast_i == 0 else "Forecast W" + str(forecast_i)


            cellValues = []
            for i, col in enumerate(pd.unique(self.df["New Routing"])):
                # sum of Stock FG Free % for each line 
                sum_line_free = self.df[(self.df["New Routing"] == col) & (self.df[forecastText] != 0)]["STOCK_FG_FREE"].sum()
                sum_forecast_w = self.df[(self.df["New Routing"] == col) & (self.df[forecastText] != 0)][forecastText].sum()
                percent = round(float(sum_line_free) / float(sum_forecast_w) * 100, 2) if sum_forecast_w != 0 else 0
                cellValues.append(float(percent))
            newOrder = [3, 4, 0, 2, 1]
            # change order of cells E64 to E68 to be in order of newOrder
            for i, newIndex in enumerate(newOrder):
                ws.cell(row=64+i, column=5 + 2*forecast_i).value = str(cellValues[newIndex]) + "%"
            ws.cell(row=69, column=5 + 2*forecast_i).value = str(round(float(sum(cellValues)/len(cellValues)), 2)) + "%"


        #Update text to include week number
        weekNumber = datetime.datetime.now(__import__("zoneinfo").ZoneInfo(CURENT_TIME_ZONE)).isocalendar()[1]

        # Week number
        for forecast_i in range(9): 
            ws.cell(row=2, column=4+2*forecast_i).value = "Forecast "+ "W" +str(weekNumber+forecast_i)
            ws.cell(row=21, column=4+2*forecast_i).value = "W"+str(weekNumber+forecast_i)+" Avance / Retard"
            ws.cell(row=29, column=4+2*forecast_i).value = "W"+str(weekNumber+forecast_i)+" Total"
            ws.cell(row=38, column=4+2*forecast_i).value = "W"+str(weekNumber+forecast_i)+" Programmation Semaine"
            ws.cell(row=53, column=4+2*forecast_i).value = "W"+str(weekNumber+forecast_i)+" Ratrappage Backlog"
            ws.cell(row=63, column=4+2*forecast_i).value = "W"+str(weekNumber+forecast_i)+" Stock FG Free %"
            ws.cell(row=72, column=4+2*forecast_i).value = "W"+str(weekNumber+forecast_i)+" Tonnes exceptionelles"



        self.summary=summary
        wb.calculation.fullCalcOnLoad = True

        # set active cell to be on sheet 1
        ws = wb["Resultats"]
        ws.sheet_view.selection[0].active_cell = "G20"
        ws.sheet_view.selection[0].sqref = "G20"

        # write to F1 the last modified date of the exception report
        if self.exceptionReportLastModified != None:
            # add one to the hour to be in the right timezone
            self.exceptionReportLastModified = self.exceptionReportLastModified
            
            lastModifiedDate = datetime.datetime.fromtimestamp(self.exceptionReportLastModified).strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=1, column=6).value = "Exception Report from : " + lastModifiedDate

            updateTime = datetime.datetime.now(__import__("zoneinfo").ZoneInfo(CURENT_TIME_ZONE)).strftime("%Y-%m-%d %H:%M:%S")
            
            ws.cell(row=1, column=8).value = "Updated at : " + updateTime

            

        printerUtil("Saving output file...")
        wb.save(template_path)
        printerUtil("Output file saved.")
        


        

    def createDetail(self, rootingIndex, protoIndex, isPostes, col, wb):
        
        
        # keep in potentialDetails the details of the lines in newDf that have the same routing and proto or not
        #printerUtil(rootingIndex, "Creating details for routing ", newLines[rootingIndex], " proto ", protoIndex, " isPostes ", isPostes, " column ", col)
        potentialDetails = self.df[self.df.apply(lambda row: (row["New Routing"] == newLines[rootingIndex]) and ((row["Is Proto"] == 'VRAI') if protoIndex == 1 else (row["Is Proto"] == 'FAUX')), axis=1)]
        
        


        
        
        tonnesColName = col.replace(" (Postes)", "") if isPostes else col 
        postesColName = col if isPostes else col + " (Postes)"


        # Go to Details sheet
        ws = wb["Details"]
        protoText = "proto" if protoIndex == 1 else "serie"
        postesText = "Postes" if isPostes else "Tonnes"

        self.start_block_row = self.curentDetailRow
        ws.cell(row=self.curentDetailRow, column=3).value = "Details for : " + newLines[rootingIndex] + " "+ protoText + " - " + postesText + " " + col.replace(" (Postes)", "")
        ws.cell(row=self.curentDetailRow, column=4).value = "Article : Client - Length x Width x Thickness" 
        ws.cell(row=self.curentDetailRow, column=5).value = "Postes" 
        ws.cell(row=self.curentDetailRow, column=6).value = "Tonnes" 
        ws.cell(row=self.curentDetailRow, column=7).value = "Productivity" 

        ws.cell(row=self.curentDetailRow+1, column=3).value = "Sum Tonnes : " + str(round(float(potentialDetails[tonnesColName].sum()), 2)) 
        ws.cell(row=self.curentDetailRow+2, column=3).value = "Sum Postes : " + str(round(float(potentialDetails[postesColName].sum()), 2))
        ws.cell(row=self.curentDetailRow+3, column=3).value = "Average Productivity : " + str(round(float(potentialDetails["Productivity"].mean()), 2))

        self.curentDetailRow=self.curentDetailRow+1


        
        realOutput = 0

        for index, row in potentialDetails.iterrows():
            #if postes + tonnes = 0 then skip the line
            if float(row[tonnesColName]) == 0 and float(row[postesColName]) == 0:
                continue

            self.productText = str(row["Material"]) + ":" +str(row["Name of sold-to party"]) + " - " + str(row["Length"]) + "x" + str(row["Width"]) + "x" + str(row["Thickness"])
            realOutput = realOutput + 1


            ws.cell(row=self.curentDetailRow, column=4).value = self.productText
            


            ws.cell(row=self.curentDetailRow, column=5).value = row[postesColName] # Postes
            ws.cell(row=self.curentDetailRow, column=6).value = row[tonnesColName] # Tonnes
            ws.cell(row=self.curentDetailRow, column=7).value = round(float(row["Productivity"]), 2) # Productivity



            self.curentDetailRow=self.curentDetailRow+1


        self.end_block_row = self.curentDetailRow - 1
        if realOutput <= 3:
            self.end_block_row = self.start_block_row + 3
            
        self.Colorizer(self.start_block_row, self.end_block_row, ws)

        self.curentDetailRow=self.curentDetailRow+3
        if realOutput <= 3:
            self.curentDetailRow=self.start_block_row + 7

                
    def Colorizer(self, start_row, end_row, ws):
        # ===== Styles =====
        header_fill = PatternFill("solid", fgColor="D9E1F2")
        summary_fill = PatternFill("solid", fgColor="F2F2F2")
        max_fill = PatternFill("solid", fgColor="C6EFCE")
        min_fill = PatternFill("solid", fgColor="FFC7CE")

        bold_font = Font(bold=True)

        thin = Side(style="thin")
        thick = Side(style="medium")

        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ===== Apply thin borders everywhere inside block =====
        for r in range(start_row, end_row + 1):
            for c in range(3, 8):
                ws.cell(row=r, column=c).border = thin_border

        # ===== Header styling =====
        for c in range(3, 8):
            cell = ws.cell(row=start_row, column=c)
            cell.fill = header_fill
            cell.font = bold_font

        # ===== Summary styling (next 3 rows in column 3 only) =====
        for r in range(start_row + 1, start_row + 4):
            ws.cell(row=r, column=3).fill = summary_fill

        # ===== Highlight Productivity max / min =====
        tonnes_values = []

        for r in range(start_row + 1, end_row + 1):
            val = ws.cell(row=r, column=6).value
            if isinstance(val, (int, float)):
                tonnes_values.append(val)

        if tonnes_values:
            max_tonnes = max(tonnes_values)
            min_tonnes = min(tonnes_values)

            for r in range(start_row + 1, end_row + 1):
                cell = ws.cell(row=r, column=6)

                if cell.value == min_tonnes:
                    # color all line in red if min_tonnes, and set font to bold
                    for c in range(4, 8):
                        ws.cell(row=r, column=c).fill = min_fill
                        ws.cell(row=r, column=c).font = bold_font

                if cell.value == max_tonnes:
                    # color all line in green if max_tonnes, and set font to bold
                    for c in range(4, 8):
                        ws.cell(row=r, column=c).fill = max_fill
                        ws.cell(row=r, column=c).font = bold_font

                

        # ===== Thick outside border =====
        for c in range(3, 8):
            ws.cell(row=start_row, column=c).border = Border(
                top=thick,
                left=thick if c == 3 else thin,
                right=thick if c == 7 else thin
            )

            ws.cell(row=end_row, column=c).border = Border(
                bottom=thick,
                left=thick if c == 3 else thin,
                right=thick if c == 7 else thin
            )

        for r in range(start_row, end_row + 1):
            ws.cell(row=r, column=3).border = Border(left=thick)
            ws.cell(row=r, column=7).border = Border(right=thick, top=thin, bottom=thin)

    
        #Add thick border on the left of column 3, and on the right of column 7, and on the top of the header, and on the bottom of the last line
        for c in range(3, 8):
            if c == 3:
                ws.cell(row=start_row, column=c).border = Border(top=thick, left=thick, right=thick)
                ws.cell(row=end_row, column=c).border = Border(bottom=thick, left=thick, right=thick)
                for row in range(start_row+1, end_row):
                    ws.cell(row=row, column=c).border = Border(left=thick, right=thick)
            elif c == 7:
                ws.cell(row=start_row, column=c).border = Border(top=thick, right=thick, bottom=thin)
                ws.cell(row=end_row, column=c).border = Border(bottom=thick, right=thick, top=thin)
            else:
                ws.cell(row=start_row, column=c).border = Border(top=thick, left=thin, right=thin)
                ws.cell(row=end_row, column=c).border = Border(bottom=thick, left=thin, right=thin)







WINDOWS_REPORT_PATH = r"Reports\Report.xlsb"
WINDOWS_ABAQUE_PATH = r"Abaque\Abaque 2025-2026.xlsx"
WINDOWS_CACHE_FILE = r"Cache\articleCachedProductivities.txt"
WINDOWS_DF_CACHE_FILE = r"Cache\Processed_Exception_report.xlsx"
WINDOWS_OUTPUT_TEMPLATE = r"OutputTemplate.xlsm"


LINUX_PREFIX = r"/home/Raftests/AMCS/bots_previsions/semaine_postes/"
LINUX_REPORT_PATH = r"Reports/Report.xlsb"
LINUX_ABAQUE_PATH = r"Abaque/Abaque 2025-2026.xlsx"
LINUX_CACHE_FILE = r"Cache/articleCachedProductivities.txt"
LINUX_DF_CACHE_FILE = r"Cache/Processed_Exception_report.xlsx"
LINUX_OUTPUT_TEMPLATE = r"OutputTemplate.xlsm"


if os.name == 'nt':  # Windows
    report_path = WINDOWS_REPORT_PATH
    abaque_path = WINDOWS_ABAQUE_PATH
    output_template = WINDOWS_OUTPUT_TEMPLATE
    df_cache_file = WINDOWS_DF_CACHE_FILE
    cache_file = WINDOWS_CACHE_FILE

    
else:  # Linux or other
    report_path = LINUX_PREFIX + LINUX_REPORT_PATH
    abaque_path = LINUX_PREFIX + LINUX_ABAQUE_PATH
    output_template = LINUX_PREFIX + LINUX_OUTPUT_TEMPLATE
    df_cache_file = LINUX_PREFIX + LINUX_DF_CACHE_FILE
    cache_file = LINUX_PREFIX + LINUX_CACHE_FILE


def printerUtil(*messages):
    timeStamp = datetime.datetime.now(__import__("zoneinfo").ZoneInfo(CURENT_TIME_ZONE)).strftime("%Y-%m-%d %H:%M:%S")
    print("[excelHandler] ", timeStamp, " - ", " ".join(str(msg) for msg in messages))


def clearCache():
    if os.path.exists(cache_file):
        os.remove(cache_file)
        printerUtil(f"Cleared: {cache_file}")
    if os.path.exists(df_cache_file):
        os.remove(df_cache_file)
        printerUtil(f"Cleared: {df_cache_file}")


def verifyReportExtention():
    global report_path

    try:
        # try to locate the report file with .xlsb extension
        if not os.path.exists(report_path):
            # change report_path to have .xlsx extension instead of .xlsb
            report_path = report_path.replace(".xlsb", ".xlsx")
            if not os.path.exists(report_path):
                printerUtil("Error: Report file not found with .xlsb or .xlsx extension.")
                exit(1)
    except Exception as e:
        printerUtil("Error while verifying report file extension: ", e)
        exit(1)



def main():
    verifyReportExtention()

    bypassCalculs = False

    if len(sys.argv) > 1:
        if sys.argv[1] == "--Bypass":
            printerUtil("Bypassing calculations and using cached data of exception report...")
            bypassCalculs = True
        elif sys.argv[1].lower() == "--clear":
            printerUtil("Clearing cache files...")
            clearCache()
        elif sys.argv[1].lower() == "--help":
            printerUtil("Usage: python excelHandler.py [OPTIONS]")
            printerUtil("\nOptions:")
            printerUtil("  --Bypass    Skip calculations and use cached data")
            printerUtil("  --clear     Clear all cache files")
            printerUtil("  --help      Show this help message")
            return
        
    printerUtil("Starting excelHandler bypass caluls: ", bypassCalculs)
    eH = excelHandler(report_path, abaque_path, "WOIPPY", bypassCalculs=bypassCalculs, cacheFile=cache_file, DF_cacheFile=df_cache_file)
    df = eH.get_newDf()
    outputFormatter(df, eH.abaqueDF, tp=output_template, exceptionReportLastModified=eH.exceptionReportLastModified)

if __name__ == "__main__":
    main()